import tkinter
import tkinter as tk
from tkinter import filedialog, INSERT
from PIL import Image, ImageTk
from file_path import mp3_path
from regex import get_bpm_key
from beatstars import beatstars
from beat_video import create_video
from tag_generator import tag_generator
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re
from tkinter import messagebox
from win10toast import ToastNotifier

class BeatUploadApp:

    def __init__(self, root):
        self.root = root
        self.root.geometry('350x580')

        self.root.title("Beat Upload Application")

        # Create Frames
        self.first_frame = tk.Frame(root)
        self.first_frame.pack()
        self.tag_frame = tk.LabelFrame(self.first_frame, text='Tag Information')
        self.tag_frame.pack(pady=(20, 0))

        self.second_frame = tk.Frame(root)
        self.second_frame.pack()
        self.number_frame = tk.LabelFrame(self.second_frame, text='Number Information')
        self.number_frame.pack(pady=(20, 0))

        self.third_frame = tk.Frame(root)
        self.third_frame.pack()
        self.path_frame = tk.LabelFrame(self.third_frame, text='Path Information')
        self.path_frame.pack(pady=(20, 0))

        self.fourth_frame = tk.Frame(root)
        self.fourth_frame.pack()
        self.check_frame = tk.LabelFrame(self.fourth_frame, text='App Checkboxes')
        self.check_frame.pack(pady=(20, 0))



        self.tags_label = tk.Label(self.tag_frame, text="Enter Tags (comma-separated):")
        self.tags_label.pack()

        self.tags_entry = tk.Entry(self.tag_frame, width=40)
        self.tags_entry.insert(INSERT, "Sexyy Red, Sukihana, Glorilla")
        self.tags_entry.pack()

        self.tags_display = tk.Label(self.tag_frame, text="", font=("Helvetica", 8, "bold"))
        self.tags_display.pack()

        self.submit_tags_button = tk.Button(self.tag_frame, text="Submit Tags", command=self.submit_tags)
        self.submit_tags_button.pack()

        self.number_label = tk.Label(self.number_frame, text="Enter Beat Number:")
        self.number_label.pack()

        self.number_entry = tk.Entry(self.number_frame)
        self.number_entry.insert(INSERT, "3330")
        self.number_entry.pack()

        self.submit_number_button = tk.Button(self.number_frame, text="Submit Number", command=self.submit_number)
        self.submit_number_button.pack()

        self.number_display = tk.Label(self.number_frame, text="", font=("Helvetica", 8, "bold"))
        self.number_display.pack()

        self.folder_label = tk.Label(self.path_frame, text="Select Folder:")
        self.folder_label.pack()

        self.folder_button = tk.Button(self.path_frame, text="Browse Folder", command=self.browser_folder)
        self.folder_button.pack()

        self.folder_display = tk.Label(self.path_frame, text="", font=("Helvetica", 8, "bold"))
        self.folder_display.pack()

        #self.upload_button = tk.Button(self.root, text="Upload Beat", command=self.upload_beat, state="disabled")
        #self.upload_button.pack(pady = (20, 0))
        self.beat_state = tk.IntVar()
        self.beat_checkbutton = tk.Checkbutton(self.check_frame, text='Upload Beat', variable=self.beat_state, state='disabled')
        self.beat_checkbutton.pack()


        #self.video_button = tk.Button(self.root, text="Create Video", command=self.create_video, state="disabled")
        #self.video_button.pack()
        self.video_state = tk.IntVar()
        self.video_checkbutton = tk.Checkbutton(self.check_frame, text='Create Video', variable=self.video_state,  state='disabled')
        self.video_checkbutton.pack()


        #self.tag_button = tk.Button(self.root, text="Update Tag Generator", command=self.update_tag_generator, state="disabled")
        #self.tag_button.pack()
        self.tag_state = tk.IntVar()
        self.tag_checkbutton = tk.Checkbutton(self.check_frame, text='Update Tag Generator', variable=self.tag_state,  state='disabled')
        self.tag_checkbutton.pack()


        self.start_button = tk.Button(self.check_frame, text="Start App", command=self.start_app)
        self.start_button.pack()

        self.cred_label = tk.Label(self.root, text="Created by https://github.com/MyNameIsCarsten", font=("Helvetica", 8), fg="grey")
        self.cred_label.pack(pady = 20)

        self.folder_path = None
        self.selected_folder = ""
        self.dir_path = None
        self.number = None  # Initialize number variable
        self.bpm = None
        self.key = None
        self.name = None
        self.tags = []

        try:
            # Try to open the file for reading
            with open("user_inputs.txt", "r") as file:
                lines = file.readlines()

            # Extract the saved values and insert them into the entry widgets
            for line in lines:
                if line.startswith("Tags:"):
                    self.tags_entry.delete(0, tk.END)
                    self.tags_entry.insert(INSERT, line.strip().replace("Tags: ", ""))
                elif line.startswith("Number:"):
                    self.number_entry.delete(0, tk.END)
                    self.number_entry.insert(INSERT, line.strip().replace("Number: ", ""))
        except FileNotFoundError:
            pass  # The file doesn't exist yet

    def submit_number(self):
        entered_number = self.number_entry.get()
        if entered_number.isdigit():
            self.number = int(entered_number)
            self.number_display.config(text=f"Submitted Number: {self.number}", fg="grey")
            self.submit_number_button.config(state="disabled")
            #self.upload_button.config(state="normal")
            #self.video_button.config(state="normal")
            #self.tag_button.config(state="normal")

        else:
            self.number_display.config(text="Invalid Input", fg="red")

    def browser_folder(self):
        self.selected_folder = filedialog.askdirectory()
        self.folder_display.config(text=f"{self.selected_folder}", fg="grey")
        self.dir_path = os.path.dirname(self.selected_folder)
        self.folder_button.config(state="disabled")
        self.beat_checkbutton.config(state="normal")
        self.tag_checkbutton.config(state="normal")
        self.video_checkbutton.config(state="normal")

    def submit_tags(self):
        entered_tags = self.tags_entry.get()
        self.tags = [tag.strip() for tag in entered_tags.split(',')]
        self.tags_display.config(text=f"{self.tags}", fg="grey")
        self.submit_tags_button.config(state="disabled")


    def upload_beat(self):
        self.folder_path = self.selected_folder.split('/')[-1]
        if self.folder_path:
            self.bpm, self.key, self.name = get_bpm_key(mp3_path(self.number, self.folder_path, self.dir_path))
            #self.tags = ["Sexyy Red", "Sukihana", "Glorilla"]  # Your tags
            driver_path = 'D:\Programme\PyCharm_Projects\chromedriver.exe'  # Your driver path

            beatstars(self.number, self.bpm, self.name, self.key, self.tags, self.folder_path, self.dir_path, driver_path)
        #self.upload_button.config(state="disabled")



    def create_video(self):
        self.folder_path = self.selected_folder.split('/')[-1]
        if self.folder_path:
            self.bpm, self.key, self.name = get_bpm_key(mp3_path(self.number, self.folder_path, self.dir_path))  # Assuming "number" is defined somewhere
            clip_path = r"D:\Dropbox\Youtube Uploads\next.png"  # Your clip path
            vid_path = fr"C:\Users\Carsten\Downloads\{self.tags[0]} - {self.name}.mp4"  # Your video path
            shorts_path = fr"D:\Dropbox\Youtube Uploads\Shorts\{self.tags[0]} - {self.name} - Short.mp4"  # Your shorts path

            create_video(mp3_path(self.number, self.folder_path, self.dir_path), clip_path, vid_path, shorts_path)
        #self.video_button.config(state="disabled")
        try:
            toaster = ToastNotifier()
            toaster.show_toast('Video is done.', 'The video was successfully created.', duration=120)
        except TypeError:
            print('Notification did not work.')


    def update_tag_generator(self):
        tag_workbook_path = r"D:\Dropbox\Youtube Uploads\Tag Generator.xlsx"

        try:
            workbook = load_workbook(filename=tag_workbook_path)
        except FileNotFoundError:
            workbook = Workbook()

        first_sheet = workbook["808 Mafia"]

        first_sheet["J3"].value = self.name
        first_sheet["D3"].value = self.tags[0]
        first_sheet["G3"].value = self.tags[1]

        workbook.save(filename=tag_workbook_path)
        #self.tag_button.config(state="disabled")
        return

    def start_app(self):
        tags_save = self.tags_entry.get()
        number_save = self.number_entry.get()

        # Save the inputs to a file
        with open("user_inputs.txt", "w") as file:
            file.write(f"Tags: {tags_save}\n")
            file.write(f"Number: {number_save}")

        if self.beat_state.get() == 1:
            self.upload_beat()
            self.beat_checkbutton.config(state="disabled")
        else:
            pass

        if self.video_state.get() == 1:
            self.create_video()
            self.video_checkbutton.config(state="disabled")
        else:
            pass

        if self.tag_state.get() == 1:
            self.update_tag_generator()
            self.tag_checkbutton.config(state="disabled")
        else:
            pass

        return

if __name__ == "__main__":
    root = tk.Tk()
    app = BeatUploadApp(root)
    root.mainloop()
